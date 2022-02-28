import warnings
from tkinter import *
import pandas as pd
from pandas import read_excel
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import os
import comtypes.client
from pandas.core.common import SettingWithCopyWarning

# You need to install the XlsxWriter module in Anaconda Prompt (pip install XlsxWriter)

# 1. Input

# 1.1 Path: Sample and Database
root = Tk()
root.geometry("200x100")
Sample_input = 0


def retrieve_sample():
    global Sample_input
    Sample_input = textBox.get("1.0", "end-1c")
    return Sample_input


textBox = Text(root, height=1, width=20)
textBox.pack()
buttonCommit = Button(root,
                      height=1,
                      width=20,
                      text="Add Sample Path",
                      command=lambda: retrieve_sample())
buttonCommit.pack()
mainloop()
try:
    Sample = read_excel(Sample_input)
except FileNotFoundError:
    print('The file is not found, please double-check the path, '
          'refer to the README when in doubt')
    exit()

root = Tk()
root.geometry("200x100")
Database_input = 0


def retrieve_database():
    global Database_input
    Database_input = textBox.get("1.0", "end-1c")
    return Database_input


textBox = Text(root, height=1, width=20)
textBox.pack()
buttonCommit = Button(root,
                      height=1,
                      width=20,
                      text="Add Database Path",
                      command=lambda: retrieve_database())
buttonCommit.pack()
mainloop()
try:
    Database = read_excel(Database_input)
except FileNotFoundError:
    print('The file is not found, please double-check the path, '
          'refer to the README when in doubt')
    exit()


# 1.2 Characteristics: Size of Mesh and Criteria
root = Tk()
root.geometry("200x100")
Mesh_input = 0.5


def retrieve_mesh():
    global Mesh_input
    Mesh_input = textBox.get("1.0", "end-1c")
    return Mesh_input


textBox = Text(root, height=1, width=20)
textBox.pack()
buttonCommit = Button(root,
                      height=1,
                      width=20,
                      text="Add Mesh",
                      command=lambda: retrieve_mesh())
buttonCommit.pack()
mainloop()

try:
    Mesh = float(Mesh_input)
except ValueError:
    print('The format is incorrect, please refer to the README and try again')
    exit()

root = Tk()
root.geometry("200x100")
Criteria_input = 1


def retrieve_criteria():
    global Criteria_input
    Criteria_input = textBox.get("1.0", "end-1c")
    return Criteria_input


textBox = Text(root, height=1, width=20)
textBox.pack()
buttonCommit = Button(root,
                      height=1,
                      width=20,
                      text="Add Criteria",
                      command=lambda: retrieve_criteria())
buttonCommit.pack()
mainloop()
try:
    Criteria = float(Criteria_input)
except ValueError:
    print('The format is incorrect, please refer to the README and try again')
    exit()

# 2. Analysis and calculation

# 2.1 Calculation % database
warnings.simplefilter(action="ignore", category=SettingWithCopyWarning)
Database['%_Data'] = ""
rows, columns = Database.shape
count = 0
total = sum(Database['#_data'])

while count < rows:
    Database['%_Data'].iloc[count] = 100*Database['#_data'].iloc[count]/total
    count = count+1

# 2.2 Calculation organisms per m2
Sample['#/m2_sam'] = ""
rows, columns = Sample.shape
count = 0

while count < rows:
    Sample['#/m2_sam'].iloc[count] = Sample['#_sam'].iloc[count]/Mesh
    count = count+1

# 2.3 Calculation % sample
total = sum(Sample['#/m2_sam'])
Sample['%_Sam'] = ""
count = 0

while count < rows:
    Sample['%_Sam'].iloc[count] = 100*Sample['#/m2_sam'].iloc[count]/total
    count = count+1

# 2.4 Calculation difference between 2.1 and 2.3
Results = Database.merge(Sample,
                         on='Species',
                         how='right')
Results['%_Diff'] = ""
count = 0

while count < rows:
    Results['%_Diff'].iloc[count] = round(Results['%_Sam'].iloc[count]
                                          - Results['%_Data'].iloc[count], 2)
    count = count + 1

# 2.5 Show species above and below criteria


def color(v):
    if v < -Criteria:
        return f"color: {'red'};"
    elif v > Criteria:
        return f"color: {'blue'};"
    else:
        return f"color: {'black'};"


Results_Fi = Results.style.highlight_null(null_color="yellow")
Results_Fin = Results_Fi.set_properties(**{'text-align': 'left',
                                           'border-color': 'Black',
                                           'border-width': 'thin',
                                           'font-size': '7px'})
Results_Final = Results_Fin.applymap(color, subset="%_Diff")

# 3. Results: Tables and Graphs

# 3.1 Table: %, % of difference and  criteria application
exporttoexcel = pd.ExcelWriter("data\\Results_Final.xlsx",
                               engine='xlsxwriter')
Results_Final.to_excel(exporttoexcel,
                       sheet_name='Results',
                       startrow=5)
exporttoexcel.save()

# 3.2 Bar-Graphs: Organisms per m2, % (database, sampling and difference)
base = np.arange(len(Results['Species']))
width = 0.10
fig, ax = plt.subplots()
ax.bar(base - width, Results['%_Data'],
       width,
       label='%_Data')
ax.bar(base, Results['%_Sam'],
       width,
       label='%_Sam')
Differencebar = ax.bar(base + width,
                       Results['%_Diff'],
                       width,
                       label='%_Diff')
ax.set_ylabel('Percentage (%)')
ax.set_xlabel('Species')
ax.grid(zorder=0)
ax.set_xticks(base, Results['Species'])
ax.set_xticklabels(Results['Species'].astype(str).values,
                   rotation='vertical')
ax.bar_label(Differencebar)
ax.set_title('Percentage (%) per Species')
ax.legend(bbox_to_anchor=(1.05, 0), loc=3, borderaxespad=0)

fig.savefig('data\\Results_Final.jpg',
            bbox_inches='tight')

# 3.3 Generation of files (Excel and pdf)
wb = openpyxl.load_workbook('data\\Results_Final.xlsx')
wb.active.add_image(openpyxl.drawing.image.Image('data\\Results_Final.jpg'), "B{:d}".format(rows+14))
wb[wb.sheetnames[0]].cell(row=2, column=4).value = 'Results table (% Difference)'
wb[wb.sheetnames[0]].cell(row=4, column=1).value = 'Below ' \
                                                   ' the table and graph of the percentage' \
                                                   ' difference between the Sample and the Database are shown.'
wb[wb.sheetnames[0]].cell(row=rows+8, column=1).value = 'Only the ' \
                                                        'species considered within the Sample are represented.'
wb[wb.sheetnames[0]].cell(row=rows+9, column=1).value = 'The boxes highlighted boxes' \
                                                        'are the species that were not in the Database.'
wb[wb.sheetnames[0]].cell(row=rows+11, column=4).value = 'Graph (% Difference)'
wb.save('data\\Results_Final.xlsx')

comtypes.client.CreateObject('Excel.Application').Visible = False
doc = comtypes.client.CreateObject('Excel.Application')\
    .Workbooks.Open(os.path.join(os.path.abspath(r'data'),
                                 'Results_Final.xlsx'))
doc.ExportAsFixedFormat(0,
                        os.path.join(os.path.abspath(r'data'),
                                     'Results_Final.pdf'))
